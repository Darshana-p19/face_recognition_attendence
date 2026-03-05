import tkinter as tk
from tkinter import messagebox, ttk
import os
import cv2
import numpy as np
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pickle

# Constants
USER_FILE = "users.xlsx"
ATTENDANCE_FILE = "attendance.xlsx"
FACES_DIR = "faces"
FEATURES_FILE = "face_features.pkl"
CAMERA_WIDTH = 640
CAMERA_HEIGHT = 480
FACE_WIDTH = 100
FACE_HEIGHT = 100
SIMILARITY_THRESHOLD = 70.0  # Higher threshold for stricter matching
MAX_RECOGNIZED_FACES = {}  # Cache to prevent repeated recognitions

# Globals
camera_running = False
video_capture = None
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
status_label = None
registered_faces = []  # List of (name, roll, features)


def setup_files_and_folders():
    """Create necessary directories and files if they don't exist"""
    os.makedirs(FACES_DIR, exist_ok=True)

    if not os.path.exists(USER_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Name", "Roll Number"])
        wb.save(USER_FILE)

    if not os.path.exists(ATTENDANCE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Roll Number", "Date", "Time"])
        wb.save(ATTENDANCE_FILE)


def load_registered_faces():
    """Load all registered faces from disk"""
    global registered_faces

    # Try to load pre-computed features
    if os.path.exists(FEATURES_FILE):
        try:
            with open(FEATURES_FILE, 'rb') as f:
                registered_faces = pickle.load(f)
            update_status(f"Loaded {len(registered_faces)} registered faces")
            return
        except Exception as e:
            update_status(f"Error loading features: {e}")

    # If features file doesn't exist or has an error, compute features from images
    registered_faces = []
    face_files = [f for f in os.listdir(FACES_DIR) if f.endswith(".jpg")]

    for file in face_files:
        parts = file.split('_')
        if len(parts) >= 2:
            roll = parts[0]
            name = parts[1].split('.')[0]

            face_path = os.path.join(FACES_DIR, file)
            face_img = cv2.imread(face_path, cv2.IMREAD_GRAYSCALE)
            if face_img is not None:
                features = extract_face_features(face_img)
                if features is not None:
                    registered_faces.append((name, roll, features))

    # Save computed features
    if registered_faces:
        try:
            with open(FEATURES_FILE, 'wb') as f:
                pickle.dump(registered_faces, f)
        except Exception as e:
            update_status(f"Error saving features: {e}")

    update_status(f"Computed features for {len(registered_faces)} faces")


def save_user_to_excel(name, roll):
    """Save user information to users.xlsx"""
    # Check if user already exists
    wb = load_workbook(USER_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == roll:  # Check roll number
            return False

    ws.append([name, roll])
    wb.save(USER_FILE)
    return True


def extract_face_features(face_img):
    """Extract more robust features from face image"""
    if face_img is None or face_img.size == 0:
        return None

    # Ensure consistent size
    face_img = cv2.resize(face_img, (FACE_WIDTH, FACE_HEIGHT))

    # Apply histogram equalization to improve contrast
    face_img = cv2.equalizeHist(face_img)

    # Extract features using Local Binary Patterns (LBP)
    # This is a simple implementation of LBP
    lbp = np.zeros_like(face_img)
    for i in range(1, face_img.shape[0] - 1):
        for j in range(1, face_img.shape[1] - 1):
            center = face_img[i, j]
            code = 0
            code |= (face_img[i - 1, j - 1] > center) << 7
            code |= (face_img[i - 1, j] > center) << 6
            code |= (face_img[i - 1, j + 1] > center) << 5
            code |= (face_img[i, j + 1] > center) << 4
            code |= (face_img[i + 1, j + 1] > center) << 3
            code |= (face_img[i + 1, j] > center) << 2
            code |= (face_img[i + 1, j - 1] > center) << 1
            code |= (face_img[i, j - 1] > center) << 0
            lbp[i, j] = code

    # Calculate LBP histogram (split image into 4x4 regions)
    features = []
    h, w = lbp.shape
    h_step, w_step = h // 4, w // 4
    for i in range(4):
        for j in range(4):
            region = lbp[i * h_step:(i + 1) * h_step, j * w_step:(j + 1) * w_step]
            hist, _ = np.histogram(region, bins=16, range=(0, 256))
            features.extend(hist)

    # Add some pixel intensity features
    for i in range(4):
        for j in range(4):
            region = face_img[i * h_step:(i + 1) * h_step, j * w_step:(j + 1) * w_step]
            avg = np.mean(region)
            std = np.std(region)
            features.extend([avg, std])

    return np.array(features, dtype=np.float32)


def compare_faces(face1_features, face2_features):
    """Compare two face feature vectors and return similarity score (0-100)"""
    if face1_features is None or face2_features is None:
        return 0

    if len(face1_features) == 0 or len(face2_features) == 0:
        return 0

    # Make sure vectors are same length by padding shorter one
    if len(face1_features) != len(face2_features):
        max_len = max(len(face1_features), len(face2_features))
        if len(face1_features) < max_len:
            face1_features = np.pad(face1_features, (0, max_len - len(face1_features)))
        if len(face2_features) < max_len:
            face2_features = np.pad(face2_features, (0, max_len - len(face2_features)))

    # Calculate histogram intersection (better for histogram features)
    min_intersect = np.sum(np.minimum(face1_features, face2_features))
    max_sum = np.sum(np.maximum(face1_features, face2_features))

    if max_sum == 0:
        return 0

    similarity = (min_intersect / max_sum) * 100
    return float(similarity)


def mark_attendance(name, roll):
    """Mark attendance in attendance.xlsx"""
    today_date = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M:%S")

    # Check if already marked
    key = f"{name}_{roll}_{today_date}"
    if key in MAX_RECOGNIZED_FACES:
        return False

    MAX_RECOGNIZED_FACES[key] = True

    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    # Avoid duplicate attendance on the same day
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == roll and row[2] == today_date:
            update_status(f"{name} already marked present today")
            return False

    ws.append([name, roll, today_date, time_now])
    wb.save(ATTENDANCE_FILE)
    update_status(f"Attendance marked for {name}")
    return True


def update_status(message):
    """Update status label with message"""
    if status_label:
        status_label.config(text=message)


def register_user():
    """Register a new user with face data"""
    global video_capture, registered_faces

    name = name_entry.get().strip()
    roll = roll_entry.get().strip()

    # Validate inputs
    if not name:
        messagebox.showerror("Error", "Name cannot be empty")
        return
    if not roll:
        messagebox.showerror("Error", "Roll Number cannot be empty")
        return

    # Make sure camera is running
    if not camera_running or not video_capture or not video_capture.isOpened():
        messagebox.showerror("Error", "Camera must be running to register")
        return

    # Capture a photo
    ret, frame = video_capture.read()
    if not ret:
        messagebox.showerror("Error", "Failed to capture image. Check camera connection.")
        return

    # Detect face
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.1, 4, minSize=(30, 30))

    if len(faces) != 1:
        messagebox.showerror("Error", "Please ensure exactly one face is visible")
        return

    # Extract and process face
    (x, y, w, h) = faces[0]
    face_img = gray[y:y + h, x:x + w]

    # Save user to Excel
    if not save_user_to_excel(name, roll):
        messagebox.showerror("Error", f"Roll number {roll} already exists")
        return

    # Extract features
    features = extract_face_features(face_img)
    if features is None:
        messagebox.showerror("Error", "Could not extract face features. Try again with better lighting.")
        return

    # Save face image
    face_path = os.path.join(FACES_DIR, f"{roll}_{name}.jpg")
    cv2.imwrite(face_path, face_img)

    # Add to registered faces
    registered_faces.append((name, roll, features))

    # Update features file
    try:
        with open(FEATURES_FILE, 'wb') as f:
            pickle.dump(registered_faces, f)
    except Exception as e:
        update_status(f"Error saving features: {e}")

    messagebox.showinfo("Success", f"User {name} registered successfully!")
    update_status(f"Registered: {name} (Roll: {roll})")
    name_entry.delete(0, tk.END)
    roll_entry.delete(0, tk.END)


def is_face_registered(face_features):
    """Check if a face is registered and return match details"""
    global registered_faces

    if not registered_faces or face_features is None:
        return None

    best_match = None
    highest_similarity = 0

    for name, roll, ref_features in registered_faces:
        similarity = compare_faces(face_features, ref_features)
        if similarity > highest_similarity:
            highest_similarity = similarity
            best_match = (name, roll, similarity)

    if best_match and best_match[2] >= SIMILARITY_THRESHOLD:
        return best_match
    return None


def recognize_faces():
    """Recognize faces in video feed and mark attendance"""
    global video_capture

    if not camera_running or not video_capture or not video_capture.isOpened():
        return

    # Capture frame
    ret, frame = video_capture.read()
    if not ret:
        update_status("Camera error: Failed to capture frame")
        return

    # Make a copy for display
    display_frame = frame.copy()

    # Convert to grayscale for face detection
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.1, 4, minSize=(30, 30))

    # Process each detected face
    for (x, y, w, h) in faces:
        # Extract face
        face_img = gray[y:y + h, x:x + w]
        features = extract_face_features(face_img)

        # Default is red for unknown/unregistered
        color = (0, 0, 255)  # BGR format (Red)
        label = "Unknown"

        # Check if face is registered
        match = is_face_registered(features)
        if match:
            name, roll, similarity = match
            label = f"{name} ({similarity:.1f}%)"
            color = (0, 255, 0)  # BGR format (Green)

            # Mark attendance
            mark_attendance(name, roll)

        # Draw rectangle around face with appropriate color
        cv2.rectangle(display_frame, (x, y), (x + w, y + h), color, 2)

        # Add name label
        y_label = y - 10 if y - 10 > 10 else y + h + 20
        cv2.putText(display_frame, label, (x, y_label),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)

    # Update video display
    display_frame_rgb = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
    img = Image.fromarray(display_frame_rgb)
    img_tk = ImageTk.PhotoImage(image=img)
    video_label.config(image=img_tk)
    video_label.image = img_tk

    # Schedule next frame processing
    video_label.after(33, recognize_faces)  # ~30 fps


def start_camera():
    """Start camera capture"""
    global video_capture, camera_running

    if camera_running:
        return

    # Load registered faces first
    load_registered_faces()

    # Try to open the camera
    video_capture = cv2.VideoCapture(0)
    if not video_capture.isOpened():
        messagebox.showerror("Error", "Could not open camera. Check connection.")
        return

    video_capture.set(3, CAMERA_WIDTH)
    video_capture.set(4, CAMERA_HEIGHT)
    camera_running = True
    start_button.config(state=tk.DISABLED)
    stop_button.config(state=tk.NORMAL)
    register_button.config(state=tk.NORMAL)
    update_status("Camera started")

    # Start face recognition
    recognize_faces()


def stop_camera():
    """Stop camera capture"""
    global video_capture, camera_running

    if not camera_running:
        return

    camera_running = False
    if video_capture:
        video_capture.release()

    video_label.config(image="")
    start_button.config(state=tk.NORMAL)
    stop_button.config(state=tk.DISABLED)
    update_status("Camera stopped")


def clear_today_attendance():
    """Clear today's attendance records"""
    global MAX_RECOGNIZED_FACES
    MAX_RECOGNIZED_FACES = {}
    update_status("Today's recognition cache cleared")


def on_closing():
    """Handle window closing"""
    stop_camera()
    root.destroy()


def view_attendance():
    """Display attendance records"""
    if not os.path.exists(ATTENDANCE_FILE):
        messagebox.showerror("Error", "No attendance records found")
        return

    # Create new window
    attend_window = tk.Toplevel(root)
    attend_window.title("Attendance Records")
    attend_window.geometry("600x400")

    # Create treeview
    tree = ttk.Treeview(attend_window)
    tree["columns"] = ("Name", "Roll", "Date", "Time")
    tree.column("#0", width=0, stretch=tk.NO)
    tree.column("Name", anchor=tk.W, width=150)
    tree.column("Roll", anchor=tk.CENTER, width=100)
    tree.column("Date", anchor=tk.CENTER, width=100)
    tree.column("Time", anchor=tk.CENTER, width=100)

    tree.heading("#0", text="")
    tree.heading("Name", text="Name")
    tree.heading("Roll", text="Roll Number")
    tree.heading("Date", text="Date")
    tree.heading("Time", text="Time")

    # Add scrollbar
    scrollbar = ttk.Scrollbar(attend_window, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)

    # Layout
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Load attendance data
    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    # Populate treeview
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if len(row) >= 4:
            tree.insert("", tk.END, iid=idx, values=(row[0], row[1], row[2], row[3]))


def adjust_threshold(value):
    """Adjust the similarity threshold"""
    global SIMILARITY_THRESHOLD
    SIMILARITY_THRESHOLD = float(value)
    threshold_label.config(text=f"Threshold: {value}%")


def reset_features():
    """Delete and recompute face features"""
    global registered_faces

    if os.path.exists(FEATURES_FILE):
        os.remove(FEATURES_FILE)

    registered_faces = []
    load_registered_faces()
    update_status("Face features reset and recomputed")


# Initialize system
setup_files_and_folders()

# Create main window
root = tk.Tk()
root.title("Face Recognition Attendance System")
root.geometry("850x650")
root.protocol("WM_DELETE_WINDOW", on_closing)

# Create frames
input_frame = tk.Frame(root, pady=10)
input_frame.pack(fill=tk.X)

button_frame = tk.Frame(root, pady=10)
button_frame.pack(fill=tk.X)

video_frame = tk.Frame(root)
video_frame.pack(fill=tk.BOTH, expand=True)

control_frame = tk.Frame(root, pady=10)
control_frame.pack(fill=tk.X)

status_frame = tk.Frame(root)
status_frame.pack(fill=tk.X, side=tk.BOTTOM)

# Input fields
tk.Label(input_frame, text="Name:").grid(row=0, column=0, padx=10, pady=5)
name_entry = tk.Entry(input_frame, width=20)
name_entry.grid(row=0, column=1, padx=10, pady=5)

tk.Label(input_frame, text="Roll Number:").grid(row=0, column=2, padx=10, pady=5)
roll_entry = tk.Entry(input_frame, width=20)
roll_entry.grid(row=0, column=3, padx=10, pady=5)

# Buttons
register_button = tk.Button(button_frame, text="Register User", command=register_user,
                            bg="#4CAF50", fg="white", width=15, state=tk.DISABLED)
register_button.grid(row=0, column=0, padx=10)

start_button = tk.Button(button_frame, text="Start Camera", command=start_camera,
                         bg="#2196F3", fg="white", width=15)
start_button.grid(row=0, column=1, padx=10)

stop_button = tk.Button(button_frame, text="Stop Camera", command=stop_camera,
                        bg="#f44336", fg="white", width=15, state=tk.DISABLED)
stop_button.grid(row=0, column=2, padx=10)

view_button = tk.Button(button_frame, text="View Attendance", command=view_attendance,
                        bg="#9C27B0", fg="white", width=15)
view_button.grid(row=0, column=3, padx=10)

# Threshold slider
threshold_label = tk.Label(control_frame, text=f"Threshold: {SIMILARITY_THRESHOLD}%")
threshold_label.pack(side=tk.TOP, pady=5)

threshold_slider = tk.Scale(control_frame, from_=50, to=95, orient=tk.HORIZONTAL,
                            command=adjust_threshold, length=300)
threshold_slider.set(SIMILARITY_THRESHOLD)
threshold_slider.pack(side=tk.TOP)

# Additional controls
control_buttons_frame = tk.Frame(control_frame)
control_buttons_frame.pack(side=tk.TOP, pady=10)

clear_button = tk.Button(control_buttons_frame, text="Clear Today's Cache",
                         command=clear_today_attendance, bg="#FF9800", fg="white", width=15)
clear_button.grid(row=0, column=0, padx=10)

reset_button = tk.Button(control_buttons_frame, text="Reset Face Features",
                         command=reset_features, bg="#607D8B", fg="white", width=15)
reset_button.grid(row=0, column=1, padx=10)

# Video display
video_label = tk.Label(video_frame, bg="black", width=CAMERA_WIDTH, height=CAMERA_HEIGHT)
video_label.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

# Status bar
status_label = tk.Label(status_frame, text="System ready", bd=1, relief=tk.SUNKEN, anchor=tk.W)
status_label.pack(fill=tk.X)

# Start the application
root.mainloop()